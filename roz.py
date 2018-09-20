import face_recognition
from face_recognition.face_recognition_cli import image_files_in_folder
from PIL import Image, ImageDraw
import argparse
import cv2
import numpy as np
import os
import random
import openpyxl
import shutil

import openface
import openface.helper
from openface.data import iterImgs

fileDir = os.path.dirname(os.path.realpath(__file__))
modelDir = os.path.join(fileDir, '', 'models')
dlibModelDir = os.path.join(modelDir, 'dlib')
openfaceModelDir = os.path.join(modelDir, 'openface')

landmarks = 'outerEyesAndNose'
dlibFacePredictor = os.path.join(dlibModelDir, "shape_predictor_68_face_landmarks.dat")
outputDir = './aligned_images/'
align = openface.AlignDlib(dlibFacePredictor)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'JPG'}


def preprocess_image(img_path):
    image = Image.open(img_path)
    if hasattr(image, '_getexif'):
        orientation = 0x0112
        exif = image._getexif()
        if exif is not None:
            try:
                orientation = exif[orientation]
                rotations = {
                    3: Image.ROTATE_180,
                    6: Image.ROTATE_270,
                    8: Image.ROTATE_90
                }
                if orientation in rotations:
                    image = image.transpose(rotations[orientation])
            except:
                print ('no orientation')
    image.save(img_path)

def preprocess_dataset(dir):
    for class_dir in os.listdir(dir):
        print(class_dir)
        # Loop through each training image for the current person
        for img_path in image_files_in_folder(os.path.join(dir, class_dir)):
            preprocess_image(img_path)

def detect_image(bgr, multi):
    rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
    if rgb is None:
        print("  + Unable to load.")
        faces = None
    else:
        if multi:
            faces = align.getAllFaceBoundingBoxes(rgb)
        else:
            faces = align.getLargestFaceBoundingBox(rgb)

    return faces

def detect(img_path, multi=True):

    landmarkMap = {
        'outerEyesAndNose': openface.AlignDlib.OUTER_EYES_AND_NOSE,
        'innerEyesAndBottomLip': openface.AlignDlib.INNER_EYES_AND_BOTTOM_LIP
    }

    landmarkIndices = landmarkMap[landmarks]


    print("=== {} ===".format(img_path))

    bgr = cv2.imread(img_path)

    return detect_image(bgr, multi)

def get_embedding(cropped_img_path):
    img = cv2.imread(cropped_img_path)
    image = face_recognition.load_image_file(cropped_img_path)
    height, width, channels = img.shape
    face_bounding_box = (0, width, height, 0)

    return face_recognition.face_encodings(image, known_face_locations=[face_bounding_box])[0]


def get_embeddings(align_dir):
    # Loop through each person in the training set

    X = []
    y = []

    for class_dir in os.listdir(outputDir):
        if not os.path.isdir(os.path.join(outputDir, class_dir)):
            continue

        print(class_dir)
        e = []
        c = 0.0

        # Loop through each training image for the current person
        for img_path in image_files_in_folder(os.path.join(outputDir, class_dir)):
            embedding = get_embedding(img_path)
            if embedding == []:
                continue

            c += 1
            if e == []:
                e = embedding
            else:
                for i in range(len(embedding)):
                    e[i] += embedding[i]

        for i in range(len(e)):
            e[i] /= c
        X.append(e)
        y.append(class_dir)

    X = np.array(X)
    y = np.array(y)

    np.save('X', X)
    np.save('y', y)

    return X, y

def show_prediction_labels_on_image(img_path, predictions):
    """
    Shows the face recognition results visually.
    :param img_path: path to image to be recognized
    :param predictions: results of the predict function
    :return:
    """
    pil_image = Image.open(img_path).convert("RGB")
    draw = ImageDraw.Draw(pil_image)

    for name, (top, right, bottom, left) in predictions:
        # Draw a box around the face using the Pillow module
        draw.rectangle(((left, top), (right, bottom)), outline=(0, 0, 255))

        # There's a bug in Pillow where it blows up with non-UTF-8 text
        # when using the default bitmap font
        name = name.encode("UTF-8")

        # Draw a label with a name below the face
        text_width, text_height = draw.textsize(name)
        draw.rectangle(((left, bottom - text_height - 10), (right, bottom)), fill=(0, 0, 255), outline=(0, 0, 255))
        draw.text((left + 6, bottom - text_height - 5), name, fill=(255, 255, 255, 255))

    # Remove the drawing library from memory as per the Pillow docs
    del draw

    # Display the resulting image
    #pil_image = pil_image.resize((700, 700))
    pil_image.show()

def test(img_path, KnownEncodings, Classes):
    image = face_recognition.load_image_file(img_path)
    f = detect(img_path)
    face_locations = []
    for ff in f:
        face_locations.append((ff.top(), ff.right(), ff.bottom(), ff.left()))

    # If no faces are found in the image, return an empty result.
    if len(face_locations) == 0:
        return []
    predictions = []
    preds = []
    # Find encodings for faces in the test iamge
    faces_encodings = face_recognition.face_encodings(image, known_face_locations=face_locations)
    for j, face_to_test in enumerate(faces_encodings):
        face_distances = face_recognition.face_distance(KnownEncodings, face_to_test)
        i = 0
        for k in range(len(face_distances)):
            if face_distances[k] < face_distances[i]:
                i = k
        if (face_distances[i] > 0.55):
            pred = "unknown"
        else:
            pred = Classes[i]
        preds.append(pred)
        predictions.append((pred, face_locations[j]))

    show_prediction_labels_on_image(img_path, predictions)
    return preds

TEMPLATE = np.float32([
    (0.0792396913815, 0.339223741112), (0.0829219487236, 0.456955367943),
    (0.0967927109165, 0.575648016728), (0.122141515615, 0.691921601066),
    (0.168687863544, 0.800341263616), (0.239789390707, 0.895732504778),
    (0.325662452515, 0.977068762493), (0.422318282013, 1.04329000149),
    (0.531777802068, 1.06080371126), (0.641296298053, 1.03981924107),
    (0.738105872266, 0.972268833998), (0.824444363295, 0.889624082279),
    (0.894792677532, 0.792494155836), (0.939395486253, 0.681546643421),
    (0.96111933829, 0.562238253072), (0.970579841181, 0.441758925744),
    (0.971193274221, 0.322118743967), (0.163846223133, 0.249151738053),
    (0.21780354657, 0.204255863861), (0.291299351124, 0.192367318323),
    (0.367460241458, 0.203582210627), (0.4392945113, 0.233135599851),
    (0.586445962425, 0.228141644834), (0.660152671635, 0.195923841854),
    (0.737466449096, 0.182360984545), (0.813236546239, 0.192828009114),
    (0.8707571886, 0.235293377042), (0.51534533827, 0.31863546193),
    (0.516221448289, 0.396200446263), (0.517118861835, 0.473797687758),
    (0.51816430343, 0.553157797772), (0.433701156035, 0.604054457668),
    (0.475501237769, 0.62076344024), (0.520712933176, 0.634268222208),
    (0.565874114041, 0.618796581487), (0.607054002672, 0.60157671656),
    (0.252418718401, 0.331052263829), (0.298663015648, 0.302646354002),
    (0.355749724218, 0.303020650651), (0.403718978315, 0.33867711083),
    (0.352507175597, 0.349987615384), (0.296791759886, 0.350478978225),
    (0.631326076346, 0.334136672344), (0.679073381078, 0.29645404267),
    (0.73597236153, 0.294721285802), (0.782865376271, 0.321305281656),
    (0.740312274764, 0.341849376713), (0.68499850091, 0.343734332172),
    (0.353167761422, 0.746189164237), (0.414587777921, 0.719053835073),
    (0.477677654595, 0.706835892494), (0.522732900812, 0.717092275768),
    (0.569832064287, 0.705414478982), (0.635195811927, 0.71565572516),
    (0.69951672331, 0.739419187253), (0.639447159575, 0.805236879972),
    (0.576410514055, 0.835436670169), (0.525398405766, 0.841706377792),
    (0.47641545769, 0.837505914975), (0.41379548902, 0.810045601727),
    (0.380084785646, 0.749979603086), (0.477955996282, 0.74513234612),
    (0.523389793327, 0.748924302636), (0.571057789237, 0.74332894691),
    (0.672409137852, 0.744177032192), (0.572539621444, 0.776609286626),
    (0.5240106503, 0.783370783245), (0.477561227414, 0.778476346951)])
TPL_MIN, TPL_MAX = np.min(TEMPLATE, axis=0), np.max(TEMPLATE, axis=0)
MINMAX_TEMPLATE = (TEMPLATE - TPL_MIN) / (TPL_MAX - TPL_MIN)

def save_aligned_images(inputDir):
    for class_dir in os.listdir(inputDir):
        if not os.path.isdir(os.path.join(inputDir, class_dir)):
            continue

        print(class_dir)
        openface.helper.mkdirP(outputDir + '/' + class_dir)
        i = 0
        for img_path in image_files_in_folder(os.path.join(inputDir, class_dir)):
            print(img_path )
            faces=detect(img_path,True)
            bgr = cv2.imread(img_path)
            rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
            landmarks = []
            for BB in faces:
                landmarks.append(align.findLandmarks(rgb, BB))
            thumbnails = []
            landmarkIndices=openface.AlignDlib.INNER_EYES_AND_BOTTOM_LIP
            for lm in landmarks:
                npLandmarks = np.float32(lm)
                npLandmarkIndices = np.array(landmarkIndices)

            H = cv2.getAffineTransform(npLandmarks[npLandmarkIndices],
                                       96 * MINMAX_TEMPLATE[npLandmarkIndices])
            thumbnail = cv2.warpAffine(rgb, H, (96, 96))
            outBgr = cv2.cvtColor(thumbnail, cv2.COLOR_RGB2BGR)
            cv2.imwrite( outputDir + class_dir + '/' + str(i) + ".png", outBgr)
            print('output' , outputDir + class_dir + '/' + str(i) + ".png")
            i += 1

    return

def RecordAttendance(WorkBookName,WorkSheetName,WeekNumber,Students):
    wb = openpyxl.load_workbook(WorkBookName)
    sheet = wb[WorkSheetName]

    for name in Students:
        for row in range(2,sheet.max_row+1):
         #Here you can add or reduce the columns
            cell_name = "{}{}".format("A", row)
            name1 = sheet[cell_name].value
            if sheet[cell_name].value == name:
                print ('got it')
                sheet.cell(row=row, column=WeekNumber + 1).value = 1
                break

    wb.save(WorkBookName)
    return

#X = np.load('X.npy')
#y = np.load('y.npy')

#save_aligned_images('./dataset')

#get_embeddings('./aligned_images')

#test('/home/higazy/Desktop/Vision pics/3.JPG', X, y)